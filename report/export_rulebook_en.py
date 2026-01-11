#!/usr/bin/env python3
import json
import sys
from pathlib import Path

import pandas as pd

try:
    import xlsxwriter  # type: ignore  # noqa: F401
    DEFAULT_ENGINE = "xlsxwriter"
    HAS_XLS_ENGINE = True
except Exception:
    try:
        from openpyxl.styles import Border, Font, PatternFill, Side  # type: ignore
        DEFAULT_ENGINE = "openpyxl"
        HAS_XLS_ENGINE = True
    except Exception:
        DEFAULT_ENGINE = None
        HAS_XLS_ENGINE = False

# Big Five mapping
TRAIT_TO_FACETS = {
    "Openness": ["Imagination", "ArtisticInterests", "Emotionality", "Adventurousness", "Intellect", "Liberalism"],
    "Conscientiousness": ["SelfEfficacy", "Orderliness", "Dutifulness", "AchievementStriving", "SelfDiscipline", "Cautiousness"],
    "Extraversion": ["Friendliness", "Gregariousness", "Assertiveness", "ActivityLevel", "ExcitementSeeking", "Cheerfulness"],
    "Agreeableness": ["Trust", "Morality", "Altruism", "Cooperation", "Modesty", "Sympathy"],
    "Neuroticism": ["Anxiety", "Anger", "Depression", "SelfConsciousness", "Immoderation", "Vulnerability"],
}

EN_COLUMNS = [
    "Marker",
    "Description",
    "Range",
    "Condition",
    "Facet",
    "Big5_Trait",
    "Weight",
    "Calibration",
]

TRAIT_BADGE = {
    "Openness": "🔵 Openness",
    "Conscientiousness": "🟠 Conscientiousness",
    "Extraversion": "🔴 Extraversion",
    "Agreeableness": "🟢 Agreeableness",
    "Neuroticism": "🟣 Neuroticism",
    "NA": "⚪ NA",
}


def facet_to_trait(facet: str) -> str:
    for trait, facets in TRAIT_TO_FACETS.items():
        if facet in facets:
            return trait
    return "NA"


def load_rulebook(path: Path) -> dict:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def build_df(rulebook: dict) -> pd.DataFrame:
    rows = []
    for marker, cfg in rulebook.items():
        desc = cfg.get("description", "")
        ranges = cfg.get("ranges", {})
        for rng, data in ranges.items():
            cond = data.get("condition", "")
            effects = data.get("facet_effects", {}) or {}
            calib = data.get("calibration", "")
            if effects:
                for facet, weight in effects.items():
                    rows.append(
                        {
                            "Marker": marker,
                            "Description": desc,
                            "Range": rng,
                            "Condition": cond,
                            "Facet": facet,
                            "Big5_Trait": TRAIT_BADGE.get(facet_to_trait(facet), "⚪ NA"),
                            "Weight": weight,
                            "Calibration": calib,
                        }
                    )
            else:
                rows.append(
                    {
                        "Marker": marker,
                        "Description": desc,
                        "Range": rng,
                        "Condition": cond,
                        "Facet": "",
                        "Big5_Trait": "",
                        "Weight": "",
                        "Calibration": calib,
                    }
                )
    return pd.DataFrame(rows, columns=EN_COLUMNS)


def merge_first_column_xlsxwriter(ws, values, fmt):
    col = 0
    n = len(values)
    i = 1  # data rows start at index 1 (0 is header)
    while i <= n:
        start = i
        while i <= n and values[i - 1] == values[start - 1]:
            i += 1
        end = i - 1
        if end > start:
            ws.merge_range(start, col, end, col, values[start - 1], fmt)
        else:
            ws.write(start, col, values[start - 1], fmt)


def merge_first_column_openpyxl(ws, values, header_border, header_fill, header_font):
    col_letter = "A"
    n = len(values)
    i = 2  # Excel rows are 1-based; row 1 is header
    while i <= n + 1:
        start = i
        while i <= n + 1 and values[i - 2] == values[start - 2]:
            i += 1
        end = i - 1
        if end > start:
            ws.merge_cells(f"{col_letter}{start}:{col_letter}{end}")
        cell = ws[f"{col_letter}{start}"]
        cell.value = values[start - 2]
        cell.fill = header_fill
        cell.font = header_font
        cell.border = header_border


def merge_column_xlsxwriter(ws, col_idx, values, fmt):
    n = len(values)
    i = 1  # data rows start at index 1 (0 is header)
    while i <= n:
        start = i
        while i <= n and values[i - 1] == values[start - 1]:
            i += 1
        end = i - 1
        if end > start:
            ws.merge_range(start, col_idx, end, col_idx, values[start - 1], fmt)
        else:
            ws.write(start, col_idx, values[start - 1], fmt)


def merge_column_openpyxl(ws, col_idx, values, border, fill, font):
    col_letter = chr(ord("A") + col_idx)
    n = len(values)
    i = 2
    while i <= n + 1:
        start = i
        while i <= n + 1 and values[i - 2] == values[start - 2]:
            i += 1
        end = i - 1
        if end > start:
            ws.merge_cells(f"{col_letter}{start}:{col_letter}{end}")
        cell = ws[f"{col_letter}{start}"]
        cell.value = values[start - 2]
        cell.fill = fill
        cell.font = font
        cell.border = border


def export_excel(df: pd.DataFrame, out_path: Path) -> None:
    if HAS_XLS_ENGINE:
        engine = DEFAULT_ENGINE
        with pd.ExcelWriter(out_path, engine=engine) as writer:
            df.to_excel(writer, index=False, sheet_name="rulebook")
            ws = writer.sheets["rulebook"]
            if engine == "xlsxwriter":
                wb = writer.book
                header_fmt = wb.add_format({"bold": True, "bg_color": "#DCE6F1", "border": 1})
                cell_fmt = wb.add_format()
                for col_num, value in enumerate(df.columns):
                    ws.write(0, col_num, value, header_fmt)
                    width = min(50, max(df[value].astype(str).map(len).max() + 2, len(value) + 2))
                    ws.set_column(col_num, col_num, width)
                # merge col 0 (Marker) and col 1 (Description)
                merge_column_xlsxwriter(ws, 0, df.iloc[:, 0].tolist(), cell_fmt)
                merge_column_xlsxwriter(ws, 1, df.iloc[:, 1].tolist(), cell_fmt)
            else:
                header_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
                header_font = Font(bold=True)
                thin = Side(border_style="thin", color="000000")
                header_border = Border(left=thin, right=thin, top=thin, bottom=thin)
                for col_num, value in enumerate(df.columns, start=1):
                    cell = ws.cell(row=1, column=col_num)
                    cell.value = value
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = header_border
                    width = min(50, max(df[value].astype(str).map(len).max() + 2, len(value) + 2))
                    ws.column_dimensions[cell.column_letter].width = width
                merge_column_openpyxl(ws, 0, df.iloc[:, 0].tolist(), header_border, header_fill, header_font)
                merge_column_openpyxl(ws, 1, df.iloc[:, 1].tolist(), header_border, header_fill, header_font)
        return
    else:
        # Fallback: write SpreadsheetML XML (Excel 2003 XML) with merged first column
        xml = ['<?xml version="1.0"?>',
               '<?mso-application progid="Excel.Sheet"?>',
               '<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"',
               ' xmlns:o="urn:schemas-microsoft-com:office:office"',
               ' xmlns:x="urn:schemas-microsoft-com:office:excel"',
               ' xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">',
               '<Styles>',
               '<Style ss:ID="Header"><Font ss:Bold="1"/><Interior ss:Color="#DCE6F1" ss:Pattern="Solid"/></Style>',
               '<Style ss:ID="DefaultCell"></Style>',
               '</Styles>',
               '<Worksheet ss:Name="rulebook">',
               '<Table>']

        # headers
        xml.append('<Row>')
        for col in df.columns:
            xml.append(f'<Cell ss:StyleID="Header"><Data ss:Type="String">{col}</Data></Cell>')
        xml.append('</Row>')

        # compute merge_down for col 0 and col 1
        def calc_merges(vals):
            n = len(vals)
            merges = [-1] * n
            i = 0
            while i < n:
                start = i
                while i < n and vals[i] == vals[start]:
                    i += 1
                merges[start] = i - start - 1
            return merges

        vals0 = df.iloc[:, 0].tolist()
        vals1 = df.iloc[:, 1].tolist()
        m0 = calc_merges(vals0)
        m1 = calc_merges(vals1)

        for idx in range(len(df)):
            xml.append('<Row>')
            for col_idx, col in enumerate(df.columns):
                val = df.iloc[idx, col_idx]
                if col_idx == 0:
                    if m0[idx] > 0:
                        xml.append(f'<Cell ss:MergeDown="{m0[idx]}" ss:StyleID="DefaultCell"><Data ss:Type="String">{val}</Data></Cell>')
                    elif m0[idx] == 0:
                        xml.append(f'<Cell ss:StyleID="DefaultCell"><Data ss:Type="String">{val}</Data></Cell>')
                    else:
                        continue
                elif col_idx == 1:
                    if m1[idx] > 0:
                        xml.append(f'<Cell ss:MergeDown="{m1[idx]}" ss:StyleID="DefaultCell"><Data ss:Type="String">{val}</Data></Cell>')
                    elif m1[idx] == 0:
                        xml.append(f'<Cell ss:StyleID="DefaultCell"><Data ss:Type="String">{val}</Data></Cell>')
                    else:
                        continue
                else:
                    v = "" if pd.isna(val) else val
                    dtype = "Number" if isinstance(v, (int, float)) else "String"
                    xml.append(f'<Cell ss:StyleID="DefaultCell"><Data ss:Type="{dtype}">{v}</Data></Cell>')
            xml.append('</Row>')

        xml.append('</Table></Worksheet></Workbook>')
        out_path.write_text("".join(xml), encoding="utf-8")


def main():
    if len(sys.argv) != 3:
        sys.exit("Usage: python export_rulebook_en.py rulebook_v2.json output.xlsx")
    in_path = Path(sys.argv[1]).expanduser()
    out_path = Path(sys.argv[2]).expanduser()
    rulebook = load_rulebook(in_path)
    df = build_df(rulebook)
    export_excel(df, out_path)
    print(f"Saved: {out_path}")


if __name__ == "__main__":
    main()
